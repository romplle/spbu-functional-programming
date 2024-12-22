import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import json

book_tree = None


def load_books(filename="4 lab/books.json"):
    with open(filename, "r", encoding="utf-8") as file:
        return json.load(file)


def create_preferences(genres, authors, keywords):
    return {
        "genres": genres,
        "authors": authors,
        "keywords": keywords
    }


def filter_books_by_genre_and_year(books, genre=None, year=None):
    return [
        book for book in books
        if (genre is None or book["genre"].lower() == genre.lower()) and
           (year is None or book["year"] >= year)
    ]


def calculate_match_score(book, preferences):
    score = 0
    if book["genre"].lower() in (genre.lower() for genre in preferences["genres"]):
        score += 10
    if any(author.lower() in (pref_author.lower() for pref_author in preferences["authors"]) for author in book["author"]):
        score += 5
    if any(keyword.strip() and keyword.lower() in book["description"].lower() for keyword in preferences["keywords"]):
        score += 2
    return score


def recommend_books(books, preferences):
    rated_books = [(book, calculate_match_score(book, preferences)) for book in books]
    return sorted(rated_books, key=lambda x: x[1], reverse=True)


def update_author_display(selected_authors, authors_text_var):
    authors_text_var.set(", ".join(selected_authors))


def update_author_suggestions(authors, search_entry, suggestions_frame, select_author_callback):
    for widget in suggestions_frame.winfo_children():
        widget.destroy()

    query = search_entry.get().strip().lower()
    if not query:
        return

    matching_authors = [author for author in authors if query in author.lower()]
    for author in matching_authors[:20]:
        btn = tk.Button(
            suggestions_frame,
            text=author,
            command=lambda a=author: select_author_callback(a),
            anchor="w",
            relief="flat",
            bg="#f0f0f0",
        )
        btn.config(width=70)
        btn.pack(fill="x", padx=5, pady=2)


def select_author(author, selected_authors, authors_text_var, search_entry, update_suggestions_callback, authors, suggestions_frame):
    if author in selected_authors:
        response = messagebox.askyesno(
            "Confirmation", f"Author '{author}' is already selected. Remove from the list?"
        )
        if response:
            selected_authors.remove(author)
    else:
        selected_authors.add(author)
    update_author_display(selected_authors, authors_text_var)
    search_entry.delete(0, tk.END)
    update_suggestions_callback(authors, search_entry, suggestions_frame, lambda a: select_author(a, selected_authors, authors_text_var, search_entry, update_suggestions_callback, authors, suggestions_frame))


def display_recommendations(recommendations, results_frame):
    global book_tree

    for widget in results_frame.winfo_children():
        widget.destroy()

    columns = ("Title", "Authors", "Year", "Genre", "Match Score")
    book_tree = ttk.Treeview(results_frame, columns=columns, show="headings", height=18)

    for col in columns:
        book_tree.heading(col, text=col)
        book_tree.column(col, anchor="w", stretch=True, width=130)

    for book, score in recommendations:
        authors = ", ".join(book["author"])
        book_tree.insert("", "end", values=(book["title"], authors, book["year"], book["genre"], score))

    scrollbar = ttk.Scrollbar(results_frame, orient="vertical", command=book_tree.yview)
    book_tree.configure(yscrollcommand=scrollbar.set)

    book_tree.grid(row=0, column=0, sticky="nsew", pady=(6, 0))
    scrollbar.grid(row=0, column=1, sticky="ns", pady=(6, 0))


def get_recommendations(books, genre_vars, selected_authors, keywords_entry, year_from_entry, year_to_entry, sort_option, sort_order, only_selected_genres_var, display_recommendations, results_frame):
    selected_genres = [genre for genre, var in genre_vars.items() if var.get()]
    preferences = create_preferences(selected_genres, list(selected_authors), keywords_entry.get().split(", "))

    filtered_books = books if not only_selected_genres_var.get() else [book for book in books if book["genre"] in selected_genres]

    year_from = year_from_entry.get().strip()
    year_to = year_to_entry.get().strip()

    if year_from.isdigit() and year_to.isdigit():
        year_from = int(year_from)
        year_to = int(year_to)
        filtered_books = [book for book in filtered_books if year_from <= book.get("year", 0) <= year_to]

    sort_reverse = (sort_order.get() == "desc")
    if sort_option.get() == "alphabet":
        filtered_books.sort(key=lambda x: x.get("title", "").lower(), reverse=sort_reverse)
    elif sort_option.get() == "year":
        filtered_books.sort(key=lambda x: x.get("year", 0), reverse=sort_reverse)

    recommendations = recommend_books(filtered_books, preferences)
    display_recommendations(recommendations, results_frame)


def save_selected_books(tree, max_col_width=70):
    if tree is None:
        messagebox.showerror("Error", "No books selected.")
        return

    selected_items = tree.selection()
    if not selected_items:
        messagebox.showwarning("Warning", "No books selected to save.")
        return

    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Books"

            headers = ["Title", "Authors", "Year", "Genre", "Match Score"]
            sheet.append(headers)

            for item in selected_items:
                values = tree.item(item, "values")
                sheet.append(values)

            for col_idx, col in enumerate(sheet.columns, start=1):
                max_length = 0
                column = get_column_letter(col_idx)
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 5, max_col_width)
                sheet.column_dimensions[column].width = adjusted_width

            workbook.save(file_path)
            messagebox.showinfo("Success", "Data saved successfully to xlsx.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save file: {e}")


root = tk.Tk()
root.title("Book Recommendation System")
root.geometry("800x600")
root.resizable(False, False)

books = load_books()
genres = sorted({book["genre"] for book in books})
authors = sorted({author for book in books for author in book["author"]})

notebook = ttk.Notebook(root)
notebook.pack(fill="both", expand=True, padx=10, pady=5)

genres_years_tab = ttk.Frame(notebook)
notebook.add(genres_years_tab, text="Genres and Years")

genres_frame = tk.LabelFrame(genres_years_tab, text="Genres", padx=10, pady=10)
genres_frame.grid(row=0, column=0, sticky="nsew")

genre_vars = {genre: tk.BooleanVar() for genre in genres}
num_columns = 4
for i, genre in enumerate(genres):
    tk.Checkbutton(genres_frame, text=genre, variable=genre_vars[genre]).grid(row=i // num_columns, column=i % num_columns, sticky="w", padx=5, pady=2)

only_selected_genres_var = tk.BooleanVar(value=False)
tk.Checkbutton(genres_frame, text="Recommend only selected genres", variable=only_selected_genres_var).grid(row=len(genres) // num_columns + 1, column=0, columnspan=num_columns, sticky="w", padx=5, pady=(0, 0))

years_frame = tk.LabelFrame(genres_years_tab, text="Publication Years", padx=5, pady=5)
years_frame.grid(row=1, column=0, sticky="nsew")

tk.Label(years_frame, text="From:").grid(row=0, column=0, padx=(10, 3), pady=(0, 4), sticky="w")
year_from_entry = tk.Entry(years_frame, width=10)
year_from_entry.grid(row=0, column=1, padx=0, pady=(0, 4))

tk.Label(years_frame, text="To:").grid(row=0, column=2, padx=(10, 3), pady=(0, 4), sticky="w")
year_to_entry = tk.Entry(years_frame, width=10)
year_to_entry.grid(row=0, column=3, padx=0, pady=(0, 4))

authors_keywords_tab = ttk.Frame(notebook)
notebook.add(authors_keywords_tab, text="Authors and Keywords")

authors_frame = tk.LabelFrame(authors_keywords_tab, text="Authors", padx=0, pady=0)
authors_frame.pack(fill="both", expand=True)

author_selection_frame = tk.Frame(authors_frame)
author_selection_frame.pack(fill="x", pady=5)

tk.Label(author_selection_frame, text="Selected authors:").pack(side="left", padx=(5, 10))
selected_authors_text = tk.StringVar(value="")
selected_authors_label = tk.Label(author_selection_frame, textvariable=selected_authors_text, wraplength=400, anchor="w", justify="left")
selected_authors_label.pack(side="left", fill="x", expand=True)

author_search_frame = tk.Frame(authors_frame)
author_search_frame.pack(fill="x", pady=5)

tk.Label(author_search_frame, text="Enter author name:").pack(side="left", padx=(5, 10))
author_search_entry = tk.Entry(author_search_frame)
author_search_entry.pack(side="left", fill="x", expand=True, padx=(5, 0))

canvas = tk.Canvas(authors_frame)
canvas.config(height=150)
scrollbar = tk.Scrollbar(authors_frame, orient="vertical", command=canvas.yview)
suggestions_frame = tk.Frame(canvas)
suggestions_frame.pack(fill="x", expand=True)

canvas.create_window((0, 0), window=suggestions_frame, anchor="nw")
canvas.config(yscrollcommand=scrollbar.set)

canvas.pack(side="left", fill="both", expand=True)
scrollbar.pack(side="right", fill="y")

suggestions_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

selected_authors = set()
author_search_entry.bind("<KeyRelease>", lambda e: update_author_suggestions(authors, author_search_entry, suggestions_frame, lambda a: select_author(a, selected_authors, selected_authors_text, author_search_entry, update_author_suggestions, authors, suggestions_frame)))

keywords_frame = tk.LabelFrame(authors_keywords_tab, text="Keywords", padx=10, pady=10)
keywords_frame.pack(fill="x")

keywords_entry = tk.Entry(keywords_frame)
keywords_entry.pack(fill="x")

recommendations_tab = ttk.Frame(notebook)
notebook.add(recommendations_tab, text="Recommendations")

results_frame = tk.LabelFrame(recommendations_tab, text="Recommendations", padx=0, pady=0)
results_frame.pack(fill="both", expand=True)

results_canvas = tk.Canvas(results_frame)
results_inner_frame = tk.Frame(results_canvas)

results_canvas.create_window((0, 0), window=results_inner_frame, anchor="nw")
results_canvas.pack(fill="both", expand=True)

sort_frame = tk.LabelFrame(recommendations_tab, text="Sorting Options", padx=10, pady=10)
sort_frame.pack(fill="both", expand=True, pady=(10, 0))

sort_option = tk.StringVar(value="alphabet")
sort_order = tk.StringVar(value="asc")

tk.Radiobutton(sort_frame, text="Alphabetically", variable=sort_option, value="alphabet").grid(row=0, column=0, sticky="w", padx=5)
tk.Radiobutton(sort_frame, text="By publication year", variable=sort_option, value="year").grid(row=0, column=1, sticky="w", padx=5)

tk.Radiobutton(sort_frame, text="Ascending", variable=sort_order, value="asc").grid(row=1, column=0, sticky="w", padx=5)
tk.Radiobutton(sort_frame, text="Descending", variable=sort_order, value="desc").grid(row=1, column=1, sticky="w", padx=5)

actions_frame = tk.Frame(root)
actions_frame.pack(fill="x", pady=10)

tk.Button(
    actions_frame,
    text="Get Recommendations",
    command=lambda: get_recommendations(
        books, genre_vars, selected_authors, keywords_entry, year_from_entry,
        year_to_entry, sort_option, sort_order, only_selected_genres_var,
        display_recommendations, results_inner_frame
    )
).pack(side="left", padx=10)

tk.Button(actions_frame, text="Save to xlsx", command=lambda: save_selected_books(book_tree)).pack(side="left", padx=10)

root.mainloop()
